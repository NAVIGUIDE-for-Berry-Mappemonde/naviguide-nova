"""
NAVIGUIDE Polar Engine
======================
- Parses polar tables from PDF or raw text
- Exports to Excel (TWA × TWS format) — raw grid OR full 180×60 extrapolated grid
- Bilinear interpolation for any TWA/TWS
- Full grid generation: raw 24×15 → extrapolated 180×60 (step 1° / 1 kt)
- VMG optimisation (upwind/downwind)

Interpolation logic (PDF → 180×60 table):
  RAW PDF data  : N_twa × N_tws sparse grid (e.g. 24 angles × 15 TWS values)
  Step 1 — bilinear interpolation between known points for every integer
            (TWA 1°…180°, TWS 1…60 kts)
  Step 2 — edge extrapolation:
            - TWA < first_known_twa  → linear fade to 0 at TWA=0°
            - TWS > 30 kts           → plateau (last known TWS column value)
  Output: numpy array shape (180, 60), indices [twa-1, tws-1]
"""

import io
import re
import math
import logging
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

log = logging.getLogger("polar_engine")

# Standard TWS columns (knots) — original sparse grid (PDF source)
STD_TWS = [0, 4, 6, 8, 10, 12, 14, 16, 20, 25, 30, 35, 40, 50, 60]

# Standard TWA rows (degrees) — original sparse grid (PDF source)
STD_TWA = [
    0, 30, 35, 40, 45, 50, 52, 60, 75, 90, 92, 102, 110, 111,
    113, 115, 120, 135, 150, 152, 160, 161, 170, 180
]

# Full output grid dimensions
FULL_TWA = list(range(0, 181))   # 0°…180° step 1° → 181 rows
FULL_TWS = list(range(0, 61))    # 0…60 kts step 1 kt → 61 columns


class PolarData:
    """Container for a boat's polar data with interpolation."""

    def __init__(self, twa_rows: List[int], tws_cols: List[int],
                 matrix: List[List[float]], boat_name: str = "Boat"):
        self.boat_name = boat_name
        self.twa_rows  = twa_rows       # list of TWA angles
        self.tws_cols  = tws_cols       # list of TWS values
        self.matrix    = np.array(matrix, dtype=float)  # shape (len_twa, len_tws)

    # ── Core interpolation ───────────────────────────────────────────────────

    def speed(self, twa: float, tws: float) -> float:
        """
        Returns boat speed (kts) at given TWA/TWS using bilinear interpolation
        on the raw sparse grid, with edge extrapolation.
        TWA in degrees (0-180), TWS in knots.
        """
        twa = abs(twa)
        twa = min(max(twa, 0.0), 180.0)
        if tws <= 0:
            return 0.0

        twa_min = self.twa_rows[0]
        tws_max = float(self.tws_cols[-1])
        eff_tws = min(tws, tws_max)

        if twa <= 0:
            return 0.0
        if twa < twa_min:
            # Linear fade from 0 at 0° to raw value at twa_min
            v_at_min = self._bilinear(float(twa_min), eff_tws)
            return round(v_at_min * (twa / twa_min), 3)
        if twa > self.twa_rows[-1]:
            twa = float(self.twa_rows[-1])

        return round(self._bilinear(twa, eff_tws), 3)

    def vmg(self, twa: float, tws: float) -> float:
        """Velocity Made Good toward the wind (positive = upwind)."""
        bs = self.speed(twa, tws)
        return bs * math.cos(math.radians(twa))

    def optimal_upwind(self, tws: float) -> Tuple[float, float, float]:
        """
        Best upwind VMG angle.
        Returns (twa_opt, boat_speed, vmg).
        """
        best_vmg, best_twa, best_bs = -999, 30, 0
        for twa in range(25, 80):
            bs  = self.speed(twa, tws)
            v   = bs * math.cos(math.radians(twa))
            if v > best_vmg:
                best_vmg, best_twa, best_bs = v, twa, bs
        return best_twa, round(best_bs, 2), round(best_vmg, 2)

    def optimal_downwind(self, tws: float) -> Tuple[float, float, float]:
        """
        Best downwind VMG angle.
        Returns (twa_opt, boat_speed, vmg_downwind).
        """
        best_vmg, best_twa, best_bs = -999, 150, 0
        for twa in range(100, 180):
            bs  = self.speed(twa, tws)
            v   = bs * math.cos(math.radians(180 - twa))   # toward leeward mark
            if v > best_vmg:
                best_vmg, best_twa, best_bs = v, twa, bs
        return best_twa, round(best_bs, 2), round(best_vmg, 2)

    def optimal_gybe_angle(self, tws: float) -> float:
        """Gybe angle = 180 - optimal_downwind_twa (symmetric)."""
        twa_opt, _, _ = self.optimal_downwind(tws)
        return round(180 - twa_opt, 1)

    # ── Summary ──────────────────────────────────────────────────────────────

    def summary(self) -> Dict:
        result = {}
        for tws in [8, 10, 12, 16, 20, 25]:
            uw_twa, uw_bs, uw_vmg = self.optimal_upwind(tws)
            dw_twa, dw_bs, dw_vmg = self.optimal_downwind(tws)
            result[tws] = {
                "upwind":   {"twa": uw_twa, "speed": uw_bs, "vmg": uw_vmg},
                "downwind": {"twa": dw_twa, "speed": dw_bs, "vmg": dw_vmg},
                "gybe_angle": self.optimal_gybe_angle(tws),
            }
        return result

    # ── Full 180×60 grid generation ──────────────────────────────────────────

    def generate_full_grid(self) -> np.ndarray:
        """
        Build the complete 181×61 matrix (TWA 0→180°, TWS 0→60 kts, step 1).

        Algorithm
        ---------
        1. Bilinear interpolation for every (twa, tws) inside the raw grid bounds.
        2. Edge extrapolation:
           - TWA < first_raw_twa  → linear fade from 0 at 0° to the raw value
           - TWA > last_raw_twa   → clamp to last row (already 180° in most polars)
           - TWS > max_raw_tws    → plateau: use speed at max_raw_tws
           - TWS = 0              → always 0

        Returns
        -------
        np.ndarray shape (181, 61): grid[twa_deg, tws_kt] = boat_speed (kts)
        """
        n_twa = len(FULL_TWA)   # 181
        n_tws = len(FULL_TWS)   # 61
        grid  = np.zeros((n_twa, n_tws), dtype=float)

        twa_min = self.twa_rows[0]
        twa_max = self.twa_rows[-1]
        tws_max = self.tws_cols[-1]

        for i, twa in enumerate(FULL_TWA):
            for j, tws in enumerate(FULL_TWS):
                if tws == 0:
                    grid[i, j] = 0.0
                    continue

                # Clamp TWS to raw maximum (plateau extrapolation)
                eff_tws = min(float(tws), float(tws_max))

                if twa <= 0:
                    grid[i, j] = 0.0

                elif twa < twa_min:
                    # Linear fade: 0 at TWA=0, raw value at twa_min
                    v_at_min = self._bilinear(float(twa_min), eff_tws)
                    grid[i, j] = v_at_min * (twa / twa_min)

                elif twa > twa_max:
                    # Clamp to last row
                    grid[i, j] = self._bilinear(float(twa_max), eff_tws)

                else:
                    grid[i, j] = self._bilinear(float(twa), eff_tws)

        log.info(f"Full grid generated: {n_twa}×{n_tws}, max={grid.max():.2f} kts")
        return grid

    def _bilinear(self, twa: float, tws: float) -> float:
        """
        Strict bilinear interpolation on the raw sparse grid.
        Both twa and tws must be within [raw_min, raw_max].
        """
        i0, i1 = self._bracket(twa, self.twa_rows)
        j0, j1 = self._bracket(tws, self.tws_cols)

        v00 = self.matrix[i0, j0]
        v10 = self.matrix[i1, j0]
        v01 = self.matrix[i0, j1]
        v11 = self.matrix[i1, j1]

        twa_span = (self.twa_rows[i1] - self.twa_rows[i0]) or 1
        tws_span = (self.tws_cols[j1] - self.tws_cols[j0]) or 1

        ft = (twa - self.twa_rows[i0]) / twa_span if i0 != i1 else 0.0
        fw = (tws - self.tws_cols[j0]) / tws_span if j0 != j1 else 0.0

        return float(
            v00 * (1-ft) * (1-fw) +
            v10 * ft     * (1-fw) +
            v01 * (1-ft) * fw     +
            v11 * ft     * fw
        )

    # ── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _bracket(val: float, arr: List) -> Tuple[int, int]:
        if val <= arr[0]:
            return 0, 0
        if val >= arr[-1]:
            return len(arr)-1, len(arr)-1
        for i in range(len(arr)-1):
            if arr[i] <= val <= arr[i+1]:
                return i, i+1
        return 0, 0


# ══════════════════════════════════════════════════════════════════════════════
# CSV / Excel Parser
# ══════════════════════════════════════════════════════════════════════════════

def parse_polar_csv(file_bytes: bytes, boat_name: str = "Boat") -> PolarData:
    """
    Parse a polar table from a CSV file.
    Expected format: first row = TWS headers, subsequent rows = TWA + speeds.
    Accepts comma, semicolon, or tab delimiters.
    """
    import pandas as pd

    # Try common delimiters
    for sep in (",", ";", "\t"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), sep=sep, header=None)
            if df.shape[1] >= 4:
                break
        except Exception:
            continue

    text = df.to_csv(sep=" ", header=False, index=False)
    log.info(f"CSV parsed: {df.shape[0]} rows × {df.shape[1]} cols")
    return parse_polar_text(text, boat_name)


def parse_polar_excel(file_bytes: bytes, boat_name: str = "Boat") -> PolarData:
    """
    Parse a polar table from an Excel file (.xlsx / .xls).
    Reads the first sheet, converts all cells to space-separated text.
    """
    import pandas as pd

    df = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=0)
    # Drop fully-empty rows/cols
    df = df.dropna(how="all").dropna(axis=1, how="all")
    # Fill NaN with 0 for numeric cells, empty string for others
    df = df.fillna("")
    text = "\n".join(
        " ".join(str(cell) for cell in row if str(cell).strip())
        for _, row in df.iterrows()
    )
    log.info(f"Excel parsed: {df.shape[0]} rows × {df.shape[1]} cols")
    return parse_polar_text(text, boat_name)


# ══════════════════════════════════════════════════════════════════════════════
# PDF Parser
# ══════════════════════════════════════════════════════════════════════════════

def parse_polar_pdf(pdf_bytes: bytes, boat_name: str = "Boat") -> PolarData:
    """
    Extract polar table from a PDF file.
    Strategy:
      1. Try extract_text() (selectable-text PDFs)
      2. Fallback to extract_tables() (vector/structured PDFs without selectable text)
    """
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber not installed")

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # ── Strategy 1: plain text extraction ─────────────────────────────────
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        text_lines = [l.strip() for l in text.splitlines() if l.strip()]
        log.info(f"extract_text() yielded {len(text_lines)} non-empty lines")

        if text_lines:
            log.info("PDF text preview (first 20 lines):\n" + "\n".join(
                f"  [{i:02d}] {l}" for i, l in enumerate(text_lines[:20])
            ))

        # ── Strategy 2: table extraction fallback ─────────────────────────────
        if not text_lines:
            log.info("No text found — attempting extract_tables() fallback")
            all_rows: List[List[str]] = []
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    for row in table:
                        cleaned = [str(cell).strip() if cell else "" for cell in row]
                        if any(cleaned):
                            all_rows.append(cleaned)

            if all_rows:
                log.info(f"extract_tables() found {len(all_rows)} rows")
                log.info("Table preview (first 20 rows):\n" + "\n".join(
                    f"  [{i:02d}] {row}" for i, row in enumerate(all_rows[:20])
                ))
                text = "\n".join(" ".join(row) for row in all_rows)
            else:
                log.warning("extract_tables() returned no data — trying OCR (Strategy 3)")

    # ── Strategy 3: OCR for raster-scan PDFs ──────────────────────────────────
    if not [l for l in text.splitlines() if l.strip()]:
        text = _ocr_pdf(pdf_bytes)

    return parse_polar_text(text, boat_name)


def _ocr_pdf(pdf_bytes: bytes) -> str:
    """
    Convert each PDF page to an image and run Tesseract OCR.
    Requires: pdf2image (poppler), pytesseract, Pillow, and tesseract binary.
    """
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError as e:
        raise RuntimeError(
            f"OCR dependencies missing ({e}). "
            "Run: pip3 install pdf2image pytesseract Pillow  "
            "and: brew install tesseract poppler"
        )

    log.info("OCR: converting PDF pages to images…")
    try:
        images = convert_from_bytes(pdf_bytes, dpi=300)
    except Exception as exc:
        raise RuntimeError(
            f"pdf2image failed: {exc}. "
            "Make sure poppler is installed: brew install poppler"
        )

    log.info(f"OCR: running Tesseract on {len(images)} page(s)…")
    pages_text = []
    for i, img in enumerate(images):
        # Use --psm 6 (uniform block of text) — best for tabular polar data
        ocr_text = pytesseract.image_to_string(
            img,
            config="--psm 6 -c tessedit_char_whitelist=0123456789. "
        )
        log.info(f"OCR page {i+1}: {len(ocr_text.splitlines())} lines extracted")
        pages_text.append(ocr_text)

    full_text = "\n".join(pages_text)
    preview = [l.strip() for l in full_text.splitlines() if l.strip()][:20]
    log.info("OCR preview (first 20 lines):\n" + "\n".join(
        f"  [{i:02d}] {l}" for i, l in enumerate(preview)
    ))
    return full_text


def _looks_like_tws_header(nums: List[float]) -> bool:
    """
    Return True if nums looks like a TWS column header.
    Criteria:
    - At least 4 values
    - All values in plausible TWS range [0, 65] kts
    - First value is a low wind speed (0–14 kts)
    - Values are mostly increasing (tolerates 1 non-monotonic step)
    """
    if len(nums) < 4:
        return False
    if any(n < 0 or n > 65 for n in nums):
        return False
    if nums[0] > 14:
        return False
    n_increasing = sum(1 for a, b in zip(nums, nums[1:]) if b >= a)
    return n_increasing >= max(len(nums) - 2, len(nums) // 2)


def parse_polar_text(text: str, boat_name: str = "Boat") -> PolarData:
    """
    Parse a polar table from raw text.
    Expected format: first non-empty row = TWS header, remaining = TWA | speeds.
    Supports Simons Voogd, PolyCurve, and generic polar PDF layouts.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # Find header line using flexible TWS detection
    tws_cols = None
    data_start = 0
    for i, line in enumerate(lines):
        nums = _extract_numbers(line)
        if _looks_like_tws_header(nums):
            tws_cols  = [int(round(n)) for n in nums]
            data_start = i + 1
            break

    if tws_cols is None:
        raise ValueError(
            "Could not detect TWS header row. "
            "Expected a row with ≥4 increasing wind speeds (0–65 kts) "
            "starting with a value ≤ 14 kts."
        )

    twa_rows = []
    matrix   = []

    for line in lines[data_start:]:
        nums = _extract_numbers(line)
        if not nums:
            continue
        twa = int(nums[0])
        if twa < 0 or twa > 180:
            continue
        speeds = nums[1:]
        # Pad or trim to match tws_cols length
        while len(speeds) < len(tws_cols):
            speeds.append(0.0)
        speeds = speeds[:len(tws_cols)]
        twa_rows.append(twa)
        matrix.append(speeds)

    if not twa_rows:
        raise ValueError("No polar data rows found.")

    log.info(f"Parsed polar: {len(twa_rows)} TWA rows × {len(tws_cols)} TWS cols")
    return PolarData(twa_rows, tws_cols, matrix, boat_name)


def _extract_numbers(line: str) -> List[float]:
    """Extract all numeric values from a line, handling decimals and comma separators."""
    # Normalize: replace commas used as decimal separators (e.g. French locale "6,5" → "6.5")
    normalized = re.sub(r"(\d),(\d)", r"\1.\2", line)
    return [float(x) for x in re.findall(r"\d+\.?\d*", normalized)]


# ══════════════════════════════════════════════════════════════════════════════
# Excel Exporter
# ══════════════════════════════════════════════════════════════════════════════

def export_to_excel(polar: PolarData) -> bytes:
    """
    Export PolarData to an Excel workbook — Berry-Mappemonde format:
    - Sheet 1: Full interpolated polar (TWA 0→180° × TWS 0→60 kts, step 1)
    - Sheet 2: Raw source data from PDF/text
    - Sheet 3: VMG optimal summary
    """
    wb = openpyxl.Workbook()

    navy   = "1F3864"
    blue   = "2E75B6"
    yellow = "FFD966"
    green  = "E2EFDA"
    white  = "FFFFFF"

    hdr_font  = Font(bold=True, color=white, name="Calibri", size=10)
    data_font = Font(name="Calibri", size=9)
    twa_font  = Font(bold=True, name="Calibri", size=9)
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Full interpolated 180×60 grid ───────────────────────────────
    ws = wb.active
    ws.title = "Polaires 180×60"

    full_grid = polar.generate_full_grid()   # shape (181, 61)
    twa_list  = FULL_TWA                     # 0…180
    tws_list  = FULL_TWS                     # 0…60

    # Title
    ws.merge_cells(f"A1:{_col(len(tws_list))}1")
    tc = ws["A1"]
    tc.value     = f"POLAIRES — {polar.boat_name.upper()}  (grille complète 181 TWA × 61 TWS, interpolation bilinéaire)"
    tc.font      = Font(bold=True, color=white, name="Calibri", size=11)
    tc.fill      = PatternFill("solid", fgColor=navy)
    tc.alignment = center

    # TWS header
    ws["A2"].value = "TWA \\ TWS"
    ws["A2"].font  = Font(bold=True, color=white, name="Calibri", size=10)
    ws["A2"].fill  = PatternFill("solid", fgColor=navy)
    ws["A2"].alignment = center

    for j, tws in enumerate(tws_list):
        c = ws.cell(row=2, column=j+2)
        c.value     = tws
        c.font      = hdr_font
        c.fill      = PatternFill("solid", fgColor=blue)
        c.alignment = center
        c.border    = border

    # Data rows — every degree of TWA
    max_speed = full_grid.max() or 1.0
    for i, twa in enumerate(twa_list):
        row      = i + 3
        is_raw   = twa in polar.twa_rows   # highlight rows present in raw data
        twa_cell = ws.cell(row=row, column=1)
        twa_cell.value     = twa
        twa_cell.font      = Font(bold=is_raw, name="Calibri", size=9,
                                  color=white if is_raw else "1F3864")
        twa_cell.fill      = PatternFill("solid", fgColor=(blue if is_raw else "D9E1F2"))
        twa_cell.alignment = center
        twa_cell.border    = border

        for j, tws in enumerate(tws_list):
            v    = full_grid[i, j]
            cell = ws.cell(row=row, column=j+2)
            cell.value     = round(v, 2) if v > 0 else ""
            cell.font      = data_font
            cell.alignment = center
            cell.border    = border
            # Green colour-scale proportional to speed
            if v > 0:
                ratio     = min(v / max_speed, 1.0)
                r_val     = int(255 - ratio * 100)
                g_val     = int(200 + ratio * 55)
                b_val     = int(220 - ratio * 80)
                hex_color = f"{r_val:02X}{g_val:02X}{b_val:02X}"
                cell.fill = PatternFill("solid", fgColor=hex_color)

    ws.column_dimensions["A"].width = 9
    for j in range(len(tws_list)):
        ws.column_dimensions[_col(j+2)].width = 5
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.freeze_panes = "B3"

    # ── Sheet 2: Raw source data ─────────────────────────────────────────────
    ws_raw = wb.create_sheet("Données brutes PDF")
    ws_raw.merge_cells(f"A1:{_col(len(polar.tws_cols))}1")
    rc = ws_raw["A1"]
    rc.value     = f"DONNÉES BRUTES — {polar.boat_name.upper()}  ({len(polar.twa_rows)} TWA × {len(polar.tws_cols)} TWS)"
    rc.font      = Font(bold=True, color=white, name="Calibri", size=11)
    rc.fill      = PatternFill("solid", fgColor=navy)
    rc.alignment = center

    ws_raw["A2"].value = "TWA \\ TWS"
    ws_raw["A2"].font  = Font(bold=True, color=white, name="Calibri", size=10)
    ws_raw["A2"].fill  = PatternFill("solid", fgColor=navy)
    ws_raw["A2"].alignment = center

    for j, tws in enumerate(polar.tws_cols):
        c = ws_raw.cell(row=2, column=j+2)
        c.value = tws; c.font = hdr_font
        c.fill  = PatternFill("solid", fgColor=blue)
        c.alignment = center; c.border = border

    for i, twa in enumerate(polar.twa_rows):
        row = i + 3
        tc2 = ws_raw.cell(row=row, column=1)
        tc2.value = twa; tc2.font = twa_font
        tc2.fill  = PatternFill("solid", fgColor=(yellow if twa in (52, 92, 111) else "D9E1F2"))
        tc2.alignment = center; tc2.border = border
        for j in range(len(polar.tws_cols)):
            v = polar.matrix[i, j]
            c = ws_raw.cell(row=row, column=j+2)
            c.value = round(v, 2) if v else ""
            c.font = data_font; c.alignment = center; c.border = border
            if v > 0:
                c.fill = PatternFill("solid", fgColor=green)

    ws_raw.column_dimensions["A"].width = 12
    for j in range(len(polar.tws_cols)):
        ws_raw.column_dimensions[_col(j+2)].width = 7

    # ── Sheet 3: VMG Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("VMG Optimal")
    headers = ["TWS (kts)", "↑ TWA opt (°)", "↑ BS (kts)", "↑ VMG (kts)",
               "↓ TWA opt (°)", "↓ BS (kts)", "↓ VMG (kts)", "Gybe angle (°)"]

    for j, h in enumerate(headers):
        c = ws2.cell(row=1, column=j+1)
        c.value     = h
        c.font      = Font(bold=True, color=white, name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=navy)
        c.alignment = center
        c.border    = border
        ws2.column_dimensions[_col(j+1)].width = 16

    summary = polar.summary()
    for row, (tws, d) in enumerate(summary.items(), start=2):
        vals = [
            tws,
            d["upwind"]["twa"],   d["upwind"]["speed"],   d["upwind"]["vmg"],
            d["downwind"]["twa"], d["downwind"]["speed"],  d["downwind"]["vmg"],
            d["gybe_angle"],
        ]
        for j, v in enumerate(vals):
            c = ws2.cell(row=row, column=j+1)
            c.value     = v
            c.font      = data_font
            c.alignment = center
            c.border    = border
            c.fill      = PatternFill("solid", fgColor=(green if row % 2 == 0 else white))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _col(n: int) -> str:
    """Convert 1-based column index to Excel letter."""
    result = ""
    while n:
        n, r = divmod(n-1, 26)
        result = chr(65+r) + result
    return result
